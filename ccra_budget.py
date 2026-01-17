#!/usr/bin/env python3
#
# openpyxl or xlswriter
#
# Alt: Data processing in python:
#   pandas library: pd.read_excel()
#   DataFrame.plot
#
# Excel functions
#   xlookup (2021), mine is 2016
#   -> in any case, would be quite clunky

import math, sys
from enum import Enum
from intervaltree import IntervalTree # pip3 install intervaltree
from openpyxl import load_workbook, Workbook # pip3 install openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# How much overrun space, in month-equivalents
SUBBUCKET_THRESHOLD = 10000
MONTH_HCELLS = 3
OVERRUN_WIDTH = 3
SQUEEZE_NARROW_BARS = True

class Err(Exception):
    def __str__(self):
        return "\n" + self.args[0]

def verbose(*args, **kwargs):
    pass
    #print(*args, **kwargs)

def account_number(s):
    match s:
        case str():
            if len(s) < 5:
                return None
            if s[4] != ' ':
                return None
            try:
                return int(s[:4])
            except:
                return None
        case int():
            return s
        case x:
            raise Err("Bad type:" + str(type(x)))

def interval_lookup(interval_tree, number):
    if number is None:
        return None
    intervals = interval_tree.at(number)
    match len(intervals):
        case 0:
            return None
        case 1:
            return next(iter(intervals)).data
        case n:
            raise

# Currently only used to check carport category
acct_categories = IntervalTree.from_tuples([
    (1000, 1999, ("Assets",)),
    (2000, 2999, ("Liabilities",)),
    (3000, 3999, ("Equities",)),
    (4000, 4199, ("Income", "Normal")),
    (4200, 4399, ("Income", "Carport")),
    (4400, 4599, ("Income", "Interest")),
    (4600, 4799, ("Income",)),
    (4800, 4899, ("Income", "Miscellaneous/bookkeeping")),
    (4900, 4999, ("Income",)),
    (5000, 5199, ("Expenses", "Operational", "Administrative")),
    (5200, 5399, ("Expenses", "Operational", "Regular services")),
    (5400, 5599, ("Expenses", "Operational", "Routine maintenance")),
    (5600, 5799, ("Expenses", "Operational", "Maintenance and repair")),
    (5800, 5999, ("Expenses", "Operational", "Miscellaneous/bookkeeping")),
    (6000, 6999, ("Expenses", "Resident Property Improvements")),
    (7000, 7999, ("Expenses", "Common Property Improvements (Capital)")),
    (8000, 8499, ("Expenses", "Carport Maintenance (Operational)")),
    (8500, 8999, ("Expenses", "Carport Improvements")),
    ])

class BucketInfo(object):
    def __init__(self, sort_id, name, account = None):
        self.sort_id = sort_id
        self.name = name
        self.account = account # if bucket lines up with a single (parent) acct

# Data is int, str pair, where int is used for sorting, 
expense_buckets = IntervalTree.from_tuples([
    (5000, 5199, BucketInfo(1, "Administrative")),
    (5200, 5399, BucketInfo(2, "Regular services")),
    (5400, 5599, BucketInfo(3, "Routine maintenance")),
    (5600, 5799, BucketInfo(4, "Repairs and Maintenance", 5600)),
    (5800, 5999, BucketInfo(5, "Contingency/Misc")),
    (6000, 6999, BucketInfo(6, "Resident Property Improvement", 6000)),
    (7000, 7999, BucketInfo(7, "Common Property Improvement", 7000)),
    (8000, 8999, BucketInfo(8, "Carport")),
    ])

# Main sections
income_range = (4000, 4999)   # lb, ub
expenses_range = (5000, 8999) # lb, ub
other_income = (4850, 4860) # Donations, Prior Period Clean Up
other_expenses = (5900,) # Asset Depreciation

def is_income(acctno):
    return (acctno is not None
            and acctno >= income_range[0] and acctno <= income_range[1]
            and acctno not in other_income)
def is_expense(acctno):
    return (acctno is not None
            and acctno >= expenses_range[0] and acctno <= expenses_range[1]
            and acctno not in other_expenses)

class Sections(object):
    def __init__(self, accts):
        self.income = []
        self.expenses = []
        self.other_income = []
        self.other_expenses = []
        self.unnumbered = [] # should not happen?
        self.other = [] # should not happen?
        self.init(accts)

    def init(self, accts):
        for acct in accts:
            acctno = account_number(acct)
            if acctno is None:
                self.unnumbered.append(acct)
                continue
            if is_income(acctno):
                self.income.append(acct)
                continue
            if is_expense(acctno):
                self.expenses.append(acct)
                continue
            if acctno in other_income:
                self.other_income.append(acct)
                continue
            if acctno in other_expenses:
                self.other_expenses.append(acct)
                continue
            self.other.append(acct)

    def __repr__(self):
        def list2str(name, lst):
            s = f"  {name}=[\n"
            for item in lst:
                s += f"    {repr(item)},\n"
            s += "  ],\n"
            return s
        return ("Sections(\n"
            + list2str("income", self.income)
            + list2str("expenses", self.expenses)
            + list2str("other_income", self.other_income)
            + list2str("other_expenses", self.other_expenses)
            + list2str("unnumbered", self.unnumbered)
            + list2str("other", self.other)
            + ")\n")

class Subaccounts(object):
    def __init__(self, tuples):
        self.num2parent = IntervalTree.from_tuples(tuples)
        self.parent_strs = set(i[2] for i in self.num2parent.items())
        self.parent_num2str = { account_number(s) :
                                s for s in self.parent_strs }

    def is_child_of(self, child, parent):
        p = interval_lookup(self.num2parent, account_number(child))
        if p is None:
            return False
        return account_number(p) == account_number(parent)

    def is_parent(self, acct):
        match acct:
            case str():
                if acct in self.parent_strs:
                    return True
                acctno = account_number(acct)
                if acctno in self.parent_num2str:
                    print("Warning: subaccount number matches but not string:",
                          acct, self.parent_num2str[acctno])
                    return True
                return False
            case int():
                return acct in self.parent_num2str
            case _:
                raise

subaccounts = Subaccounts([
    #(4401, 4499, "4400 Interest Income"), # Not subaccounts - better this way
    (5601, 5799, "5600 Repairs and Maintenance"),
    (6001, 6999, "6000 Resident Property Improvement"), # FIXME ment/ments
    (7001, 7999, "7000 Common Property Improvements"), # FIXME ment/ments
    ])

class Acct2Entries(object):
    def __init__(self):
        self.numbered = dict()
        self.unnumbered = dict()
        self.belowline = dict()
        self.acct_header = None
        self.headers = set()
        self.prefix_rows = []
        self.suffix_rows = []
        self.months = None  # str "January" or "January-April"
        self.nmonths = None # int
        # These conceptually belong to a different class;
        # stashing them here for convenience
        self.carport_income_rownos = []
        self.carport_expense_rownos = []

    def set_months(self, months, nmonths):
        assert self.months is None
        assert self.nmonths is None
        self.months = months
        self.nmonths = nmonths

    def collect_carport_rows(self, acct, rowno):
        acctno = account_number(acct)
        if acctno is None:
            print("Warning: carport check with no acctno:", acct)
            return
        ac = interval_lookup(acct_categories, acctno)
        if ac is None:
            print(f"Warning: category look failed for {acctno} ({acct})")
            return
        if len(ac) >= 2 and ac[1].startswith("Carport"):
            if ac[0] == "Income":
                self.carport_income_rownos.append(rowno)
            elif ac[0] == "Expenses":
                self.carport_expense_rownos.append(rowno)
            else:
                print(f"Warning: unexpected carport category {ac[0]} ({acct})")

    def numbered_get_entries(self, acct):
        if not acct in self.numbered:
            self.numbered[acct] = dict()
        return self.numbered[acct]

    def unnumbered_get_entries(self, acct):
        if not acct in self.unnumbered:
            self.unnumbered[acct] = dict()
        return self.unnumbered[acct]

    def belowline_get_entries(self, acct):
        if not acct in self.belowline:
            self.belowline[acct] = dict()
        return self.belowline[acct]

    def __repr__(self):
        rv = "Acct2Entries(\n"
        rv += f"  Acct Header: {self.acct_header}\n"
        rv += f"  Headers({len(self.headers)}):"
        for h in self.headers:
            rv += f" {h}"
        rv += f"\n  Prefix Rows({len(self.prefix_rows)}):\n"
        for r in self.prefix_rows:
            rv += f"    {r}\n"
        rv += f"\n  Suffix Rows({len(self.suffix_rows)}):\n"
        for r in self.suffix_rows:
            rv += f"    {r}\n"
        rv += f"\n  Numbered Accts({len(self.numbered)}):\n"
        for k1 in sorted(self.numbered.keys()):
            rv += f"    {k1}\n"
            for k2 in sorted(self.numbered[k1].keys()):
                rv += f"      {k2}={self.numbered[k1][k2]}\n"
        rv += f"\n  Unnumbered Accts({len(self.unnumbered)}):\n"
        for k1 in sorted(self.unnumbered.keys()):
            rv += f"    {k1}\n"
            for k2 in sorted(self.unnumbered[k1].keys()):
                rv += f"      {k2}={self.unnumbered[k1][k2]}\n"
        rv += f"\n  Below-the-line({len(self.belowline)}):\n"
        for k1 in sorted(self.belowline.keys()):
            rv += f"    {k1}\n"
            for k2 in sorted(self.belowline[k1].keys()):
                rv += f"      {k2}={self.belowline[k1][k2]}\n"
        rv += ")"
        return rv


def is_belowline(a2e, rowvals):
    if rowvals[0].startswith("Carport reserve"):
        return True
    if rowvals[0].startswith("Capital reserve"):
        return True
    if rowvals[0].startswith("2210 Tompkins - Line of Credit"):
        return True
    if rowvals[0] == "Net Adjusted Income":
        return True
    return False


def add_entries(entries, entry_headers, rowvals):
    if len(rowvals) != len(entry_headers) + 1:
        ncols = len(rowvals) - 1
        nhdrs = len(entry_headers)
        print("Warning: expected {nhdrs} got {ncols} columns, at {rowvals}")
        assert ncols > nhdrs
    for i, hdr in enumerate(entry_headers):
        if hdr in entries:
            print(f"Warning: duplicate entry, hdr={hdr} acct={acct}")
        entries[hdr] = rowvals[i+1]

class ReadMode(Enum):
    PREFIX = 1
    TABLE = 2
    BELOWLINE = 3
    SUFFIX = 4

# TODO "Distribution account" :
# TODO "Vendor" :

def read_entries(ws, a2e,
                 table_header1 = "Distribution account",
                 keep_affixes = False):
    # Prefix -> Table -> Belowline -> Suffix
    readmode =  ReadMode.PREFIX
    entry_headers = None
    for row in ws.rows:
        rowvals = [x.value for x in row]
        acct = rowvals[0]
        if readmode == ReadMode.PREFIX:
            assert entry_headers is None
            if acct != table_header1:
                # Haven't found headers: this is a prefix
                if keep_affixes:
                    a2e.prefix_rows.append(rowvals)
                continue
            if a2e.acct_header and a2e.acct_header != table_header1:
                print("Warning: header mismatch",
                      a2e.acct_header, table_header1)
            a2e.acct_header = table_header1
            entry_headers = rowvals[1:]
            a2e.headers.update(entry_headers)
            readmode = ReadMode.TABLE
            continue
        if readmode == ReadMode.TABLE:
            if not acct:
                readmode = ReadMode.BELOWLINE
                continue
            if account_number(acct):
                entries = a2e.numbered_get_entries(acct)
            else:
                entries = a2e.unnumbered_get_entries(acct)
            add_entries(entries, entry_headers, rowvals)
            continue
        if readmode == ReadMode.BELOWLINE:
            if not acct:
                readmode = ReadMode.SUFFIX
                continue
            if is_belowline(a2e, rowvals):
                entries = a2e.belowline_get_entries(acct)
                add_entries(entries, entry_headers, rowvals)
                continue
            else:
                readmode = ReadMode.SUFFIX
            # fallthru...
        assert readmode == ReadMode.SUFFIX
        if keep_affixes:
            a2e.suffix_rows.append(rowvals)
    return a2e

def check_header_presence(ref, data):
    for hdr in ref:
        if hdr not in data:
            print("Warning: {hdr} absent in data")

def arial(cell, fontsize = 8, bold = False, italic = False, color = None):
    cell.font = Font(name="Arial", size=fontsize,
                     bold=bold, italic=italic, color=color)

class Style(object):
    def __init__(self, bold = False, italic = False, fontsize = 8, indent = 0,
                 underline = False, overline = False, center = False):
        self.bold = bold
        self.italic = italic
        self.fontsize = fontsize
        self.indent = indent
        self.underline = underline
        self.overline = overline
        self.center = center

    def _font(self, cell):
        arial(cell, self.fontsize, self.bold, self.italic)
    def _border(self, cell, underline, overline):
        if not underline and not overline:
            return
        top = None
        if overline:
            top = Side(border_style = 'thin', color = 'FF000000')
        bottom = None
        if underline:
            bottom = Side(border_style = 'thin', color = 'FF000000')
        cell.border = Border(top = top, bottom = bottom)
    def _center(self):
        return 'center' if self.center else None
    def text(self, cell): # For headers and footers
        self._font(cell)
        cell.alignment = Alignment(horizontal = self._center(), vertical='top')
    def acct(self, cell): # For account column entry (or header)
        self._font(cell)
        self._border(cell, self.underline, False) # no overline for account
        cell.alignment = Alignment(horizontal = self._center(), vertical='top',
                                   indent=self.indent, wrap_text = True)
    def actual(self, cell): # For actual column entry (or header) ~= budget
        self._font(cell)
        self._border(cell, self.underline, self.overline)
        cell.alignment = Alignment(horizontal = self._center(), vertical='top')
        cell.number_format = "#,##0.00"
    def budget(self, cell): # For budget column entry (or header) ~= actual
        self._font(cell)
        self._border(cell, self.underline, self.overline)
        cell.alignment = Alignment(horizontal = self._center(), vertical='top')
        cell.number_format = "#,##0.00"
    def pct(self, cell): # For pct column entry (or header)
        self._font(cell)
        self._border(cell, self.underline, self.overline)
        cell.alignment = Alignment(horizontal = self._center(), vertical='top')
        cell.number_format = "##0.0%"
    def note(self, cell): # For note column entry (or header)
        self._font(cell)
        self._border(cell, self.underline, False) # no overline for note
        cell.alignment = Alignment(horizontal = self._center(), vertical='top',
                                   wrap_text = True)

def set_column_widths(ws):
    # Set column widths - these work well for 2025
    ws.column_dimensions['A'].width = 31 # =310px
    ws.column_dimensions['B'].width = 9
    ws.column_dimensions['C'].width = 9
    ws.column_dimensions['D'].width = 7
    ws.column_dimensions['E'].width = 1
    ws.column_dimensions['F'].width = 26

def add_row(ws, rowno, acct,
            actual = None, budget = None, notes = None, pct = None,
            style = Style()):
    a = ws.cell(row = rowno, column = 1, value = acct)   # A
    style.acct(a)
    b = ws.cell(row = rowno, column = 2, value = actual) # B
    style.actual(b)
    c = ws.cell(row = rowno, column = 3, value = budget) # C
    style.budget(c)
    if pct is None and budget:
        pct = f"=B{rowno}/C{rowno}"
    d = ws.cell(row = rowno, column = 4, value = pct)    # D
    style.pct(d)
    f = ws.cell(row = rowno, column = 6, value = notes)  # F
    style.note(f)
    return (a, b, c, d, f)

# Compare ActualBudgetOverrun below
class ActualBudget(object):
    def __init__(self, actual = 0, budget = 0):
        self.actual = actual
        self.budget = budget
    def __bool__(self):
        return bool(self.actual or self.budget)
    def __add__(self, rhs):
        return ActualBudget(round(self.actual + rhs.actual, 2),
                            round(self.budget + rhs.budget, 2))
    def __sub__(self, rhs):
        return ActualBudget(round(self.actual - rhs.actual, 2),
                            round(self.budget - rhs.budget, 2))
    def __repr__(self):
        return f"actual={self.actual} budget={self.budget}"

def sanity_check_numbers(a2e, acct, total):
    entry = a2e.unnumbered.get(acct, None)
    if entry is None:
        return
    # Each may be None, float, or formula string (like "=B7+B8...")
    actual = entry.get("Total", None)
    budget = entry.get("Budget", None)
    if isinstance(actual, float) and not math.isclose(actual, total.actual):
        print(f"Warning: {acct} actual mismatch: {actual} vs {total.actual}")
    if isinstance(budget, float) and not math.isclose(budget, total.budget):
        print(f"Warning: {acct} budget mismatch: {budget} vs {total.budget}")

def addcells_section(ws, rowno, secname, accts, a2e):
    add_row(ws, rowno, secname)
    rowno += 1
    total = ActualBudget()
    curparent = None
    curparent_tot = None # tied to curparent
    for acct in accts:
        entries = a2e.numbered[acct]
        actual = entries.get("Total", None)
        budget = entries.get("Budget", None)
        notes = entries.get("Notes", None)
        indent = 1
        if curparent is not None:
            if subaccounts.is_child_of(acct, curparent):
                verbose("  Child", acct)
                indent = 2
                curparent_tot.actual += 0 if actual is None else actual
                curparent_tot.budget += 0 if budget is None else budget
            else:
                # Close out curparent and curparent_tot
                verbose("End parent", acct)
                totalname = "Total for " + curparent
                add_row(ws, rowno, totalname,
                        curparent_tot.actual, curparent_tot.budget,
                        style = Style(bold = True, indent = 1))
                rowno += 1
                # Sanity check against a2e entry
                sanity_check_numbers(a2e, totalname, curparent_tot)
                curparent = None
                curparent_tot = None
        elif subaccounts.is_parent(acct):
            if actual == 0:
                actual = None
            if budget == 0:
                budget = None
            curparent = acct
            curparent_tot = ActualBudget()
            verbose("Begin parent", acct)
        a2e.collect_carport_rows(acct, rowno)
        add_row(ws, rowno, acct, actual, budget, notes,
                style = Style(indent = indent))
        rowno += 1
        total.actual += 0 if actual is None else actual
        total.budget += 0 if budget is None else budget
    totalname = "Total for " + secname
    if accts:
        add_row(ws, rowno, totalname, total.actual, total.budget,
                style = Style(bold = 1, overline = 1))
        rowno += 1
    # Sanity check against a2e entry
    sanity_check_numbers(a2e, totalname, total)
    return rowno, total

def merge_row(ws, rowno, lcol = 1, rcol = 6):
    ws.merge_cells(start_row=rowno, end_row=rowno,
                   start_column=lcol, end_column=rcol)

MONTHS = {
    "January"   : 1,
    "February"  : 2,
    "March"     : 3,
    "April"     : 4,
    "May"       : 5,
    "June"      : 6,
    "July"      : 7,
    "August"    : 8,
    "September" : 9,
    "October"  : 10,
    "November" : 11,
    "December" : 12,
}

def parse_duration(s):
    # "January 2026"
    if s.split(" ", 1)[0] == "January":
        return 1
    # "January-February, 2026"
    months = s.split(", ", 1)[0].split("-", 1)
    month1 = MONTHS.get(months[0], None)
    if month1 is None:
        return None
    if len(months) == 1:
        print("Duration is one month:", s)
        return 1
    month2 = MONTHS.get(months[1], None)
    if month2 is None:
        print("Warning: expecting second month:", s, months)
    assert month2 > month1
    return month2 - month1 + 1
    
def addcells_pnl_vs_budget(ws, pnlws, a2e):
    rowno = 1
    ###########
    # Prefix Rows
    for row in a2e.prefix_rows:
        style = None
        fontsize = { 1: 14, 2: 12, 3: 10 }.get(rowno, None)
        if fontsize is not None:
            style = Style(fontsize=fontsize, bold=True, center=True)
        for colno, val in enumerate(row, start = 1):
            if not val:
                continue
            if val == "Profit and Loss":
                val += " vs Budget"
            dur = parse_duration(val)
            if dur is not None:
                a2e.set_months(val, dur)
                val += " (" + str(round(dur * 100 / 12)) + "% of year)"
            newcell = ws.cell(row = rowno, column = colno, value = val)
            if style is not None:
                style.text(newcell)
            merge_row(ws, rowno)
        rowno += 1
    ###########
    # Headers
    verbose("Headers:", a2e.headers)
    check_header_presence(("Total", "Budget", "Notes"), a2e.headers)
    add_row(ws, rowno, a2e.acct_header,
            "Actual", "Budget", "Notes", pct = "Pct",
            style = Style(bold = True, fontsize = 9,
                          underline = True, center = True))
    rowno += 1
    ###########
    # Table
    secs = Sections(sorted(a2e.numbered))
    if secs.unnumbered:
        print(f"Warning: unnumbered accounts ({len(secs.unnumbered)}):")
        for acct in secs.unnumbered:
            print("  ", acct)
    if secs.other:
        print(f"Warning: unexpected accounts ({len(secs.other)}):")
        for acct in secs.other:
            print("  ", acct)
    # Income and Expenses
    rowno, income = addcells_section(
                        ws, rowno, "Income", secs.income, a2e)
    rowno, expenses = addcells_section(
                        ws, rowno, "Expenses", secs.expenses, a2e)
    # Net Operating Income
    netop = income - expenses
    nopi = "Net Operating Income"
    add_row(ws, rowno, nopi, netop.actual, netop.budget,
            style = Style(bold = True, overline = True))
    rowno += 1
    sanity_check_numbers(a2e, nopi, netop)
    # Other Income and Expenses
    rowno, oincome = addcells_section(
                        ws, rowno, "Other Income", secs.other_income, a2e)
    rowno, oexpenses = addcells_section(
                        ws, rowno, "Other Expenses", secs.other_expenses, a2e)
    # Net Other Income
    netoth = oincome - oexpenses
    noti = "Net Other Income"
    add_row(ws, rowno, noti, netoth.actual, netoth.budget,
            style = Style(bold = True, overline = True))
    rowno += 1
    sanity_check_numbers(a2e, noti, netoth)
    # Net Income
    netinc = netop + netoth
    ni = "Net Income"
    net_income_rowno = rowno
    add_row(ws, rowno, ni, netinc.actual, netinc.budget,
            style = Style(bold = True, overline = True))
    rowno += 1
    sanity_check_numbers(a2e, ni, netinc)
    ###########
    # Below the line
    # TODO: alternatively, incorporate this into Sections above?
    rowno += 1
    btl_rownos = []
    for acct in a2e.belowline: # Note: dict insertion order is preserved!
        entries = a2e.belowline[acct]
        actual = entries.get("Total", None)
        budget = entries.get("Budget", None)
        notes = entries.get("Notes", None)
        if acct.startswith("Carport reserve"):
            assert actual is None
            inc = "+".join(f"B{rno}" for rno in a2e.carport_income_rownos)
            exp = "+".join(f"B{rno}" for rno in a2e.carport_expense_rownos)
            assert inc
            assert exp
            actual = f"=({exp})-({inc})"
        if acct == "Net Adjusted Income":
            assert actual is None
            assert btl_rownos
            actual = f"=B{net_income_rowno}+" + "+".join(
                f"B{rno}" for rno in btl_rownos)
            style = Style(bold = True, italic = True, overline = True)
        else:
            style = Style(italic = True)
        btl_rownos.append(rowno)
        cells = add_row(ws, rowno, acct, actual, budget, notes, style = style)
        cells[0].fill = PatternFill(patternType='solid', fgColor='DDEBF7')
        rowno += 1
    rowno += 1
    ###########
    # Suffix Rows
    for row in a2e.suffix_rows:
        nonempty = []
        for colno, val in enumerate(row, start = 1):
            if val is not None:
                newcell = ws.cell(row = rowno, column = colno, value = val)
                Style().text(newcell)
                nonempty.append(colno)
        # This is for footer line ("Accrual Basis...")
        if nonempty == [1]:
            merge_row(ws, rowno)
        rowno += 1
    ###########

# Compare ActualBudget above
class ActualBudgetOverrun(object):
    def __init__(self, actual = 0, budget = 0, proj_ovr = 0):
        self.actual = actual
        self.budget = budget
        self.proj_ovr = proj_ovr
    def __bool__(self):
        return bool(self.actual or self.budget or self.proj_ovr)
    def __add__(self, rhs):
        return ActualBudgetOverrun(
            self.actual + rhs.actual,
            self.budget + rhs.budget,
            self.proj_ovr + rhs.proj_ovr)
    def __sub__(self, rhs):
        return ActualBudgetOverrun(
            self.actual - rhs.actual,
            self.budget - rhs.budget,
            self.proj_ovr - rhs.proj_ovr)
    def __repr__(self):
        return (f"actual={self.actual} budget={self.budget}"
                + (f" proj_ovr={self.proj_ovr}" if self.proj_ovr else ""))

class BucketData(object):
    def __init__(self, name):
        self.name = name
        self.absorbed_accts = set() # set(str)
        self.abo = ActualBudgetOverrun() # alt: inherit?
        self.subbuckets = dict() # account:str -> ActualBudgetOverrun()

class GraphParams(object):
    def __init__(self, sheetname, months, nmonths,
                 subbucket_threshold = SUBBUCKET_THRESHOLD,
                 month_hcells = MONTH_HCELLS,
                 overrun_width = OVERRUN_WIDTH, # in months
                 cell_width = 0.7,
                 vpxl_amt = 1000,
                 cell_minheight = 10):
        self.sheetname = sheetname # Name of worksheet
        self.months = months # String of the form "January-May"
        self.nmonths = nmonths # Number of months covered
        self.subbucket_threshold = subbucket_threshold # Min budget to separate
        self.month_hcells = month_hcells # #horizontal cells per month
        self.overrun_width = overrun_width # overrun space, in months
        self.cell_width = cell_width
        self.vpxl_amt = vpxl_amt # Dollar amount per vertical pixel
        self.cell_minheight = cell_minheight
        # Style "constants"
        self.green = PatternFill(patternType='solid', fgColor='00CC00')
        self.cyan = PatternFill(patternType='solid', fgColor='00CCCC')
        self.yellow = PatternFill(patternType='solid', fgColor='FFFFCC')
        self.redshaded = PatternFill(patternType='gray125', fgColor='FF0000')
        self.red1 = PatternFill(patternType='solid', fgColor='FF9999')
        self.red2 = PatternFill(patternType='solid', fgColor='FF0000')
        self.nofill = PatternFill(patternType=None)
        blackline = Side(border_style='thin', color='FF000000')
        bluedash = Side(border_style='dashed', color='FF0000FF')
        reddash = Side(border_style='dashed', color='FFFF0000')
        self.black_tb = Border(top=blackline, bottom=blackline)
        self.black_tbl = Border(
            top=blackline, bottom=blackline, left=blackline)
        self.black_tbr = Border(
            top=blackline, bottom=blackline, right=blackline)
        self.black_tblr = Border(
            top=blackline, bottom=blackline, left=blackline, right=blackline)
        self.bluedash_r = Border(right=bluedash)
        self.black_tb_bluedash_r = Border(
            top=blackline, bottom=blackline, right=bluedash)
        self.reddash_tb = Border(top=reddash, bottom=reddash)
        self.reddash_tbr = Border(
            top=reddash, bottom=reddash, right=reddash)
        self.reddash_tblr = Border(
            top=reddash, bottom=reddash, left=reddash, right=reddash)
        self.ROW_0 = 4
        self.COL_0 = 2
        self.COL_N = self.COL_0 + self.month_hcells * 12
        # extra 'month' for overrun:
        self.COL_OVR = self.COL_N + self.overrun_ncells()
        # Numbers and descriptions
        self.COL_PCT = self.COL_OVR + 1
        self.COL_ACTBUD = self.COL_OVR + 2
        self.COL_PRJOVR = self.COL_OVR + 3
        self.COL_DESCR = self.COL_OVR + 4
    def horiz_ncells(self):
        return 12 * self.month_hcells
    def overrun_ncells(self):
        return self.overrun_width * self.month_hcells
    def minheight_dollar(self):
        return self.vpxl_amt * self.cell_minheight

def normalize_expense_buckets(buckets, minheight_dollar):
    """
    If bucket value is small and has one subbucket, re-absorb subbucket.
    if parent bucket has one absorbed acct, push down to that acct.
    """
    for bucket_id, bd in buckets.items():
        # If bucket has one subbucket and its budget/actual is small,
        # re-absorb the subbucket
        if (len(bd.subbuckets) == 1 and
            max(bd.abo.budget + bd.abo.proj_ovr, bd.abo.actual)
                < minheight_dollar
            ):
            verbose("Reabsorbing", bd.subbuckets.keys(), "into bucket", bd.name)
            for acct, abo in bd.subbuckets.items():
                bd.absorbed_accts.add(acct)
                bd.abo += abo
            bd.subbuckets.clear()
        # If bucket has one absorbed account, push data down
        if len(bd.absorbed_accts) == 1:
            acct = next(iter(bd.absorbed_accts))
            verbose("Pushing singleton bucket", bd.name, "to", acct)
            assert not acct in bd.subbuckets
            bd.subbuckets[acct] = bd.abo
            bd.abo = ActualBudgetOverrun()

def collect_expense_buckets(a2e, gp) -> dict[int, BucketData]:
    buckets = dict()
    for acct in a2e.numbered:
        acctno = account_number(acct)
        bi = interval_lookup(expense_buckets, acctno)
        if bi is None:
            continue # non-expense
        bd = buckets.setdefault(bi.sort_id, BucketData(bi.name))
        entries = a2e.numbered[acct]
        actual = entries.get("Total", 0)
        budget = entries.get("Budget", 0) or 0
        proj_ovr = entries.get("Proj Overrun", 0) or 0
        if not actual and not budget:
            assert not proj_ovr
            continue # Skip empty rows: should just be parent accounts
        if not budget:
            # NOTE: for unbudgeted rows, Proj Overrun is total estimated cost.
            # We graph this as actual + remainder.
            if proj_ovr == 0:
                pass
            elif proj_ovr < actual:
                print(f"Warning {acct}: proj_ovr={proj_ovr} < actual={actual}")
                proj_ovr = 0
            else:
                proj_ovr -= actual
        abo = ActualBudgetOverrun(actual, budget, proj_ovr)
        if budget < gp.subbucket_threshold:
            bd.absorbed_accts.add(acct)
            bd.abo += abo
        else:
            assert not acct in bd.subbuckets
            bd.subbuckets[acct] = abo
    normalize_expense_buckets(buckets, gp.minheight_dollar())
    return buckets

def round_divide_min1(num, denom, descr):
    rv = round(num / denom, 1)
    if rv < 1:
        print(f"Warning: rounding {descr} up to one: {num}/{denom}")
        return 1
    return rv

def render_bucket(ws, rowno, gp, descr, abo, indent = False):
    curmonth_idx = None
    if gp.nmonths is not None:
        curmonth_idx = gp.nmonths * gp.month_hcells - 1
    nmonths = gp.nmonths if gp.nmonths is not None else 0
    month_hcells = gp.month_hcells
    bar_left = 0
    if not abo.budget:
        width = gp.horiz_ncells()
        po_width = round(width * abo.proj_ovr / abo.actual)
        height = round_divide_min1(abo.actual, gp.vpxl_amt, "actual height")
        if height < gp.cell_minheight:
            width = round(width * height / gp.cell_minheight)
            po_width = round(po_width * height / gp.cell_minheight)
            height = gp.cell_minheight
        bar_right = actual_width = width
        spent_color = gp.cyan
        unspent_color = gp.nofill
    else:
        bar_right = gp.horiz_ncells()
        height = round_divide_min1(abo.budget, gp.vpxl_amt, "budget height")
        if SQUEEZE_NARROW_BARS: # if line is too thin, squeeze in both sides
            while month_hcells > 1 and height < gp.cell_minheight:
                month_hcells -= 1
                height *= 2
                bar_right -= 12 - nmonths
                bar_left += nmonths
            if bar_left:
                print("Squeezing to", (bar_left, bar_right), ":", descr)
        actual_width = round(abo.actual * (bar_right - bar_left) / abo.budget)
        po_width = round(abo.proj_ovr * (bar_right - bar_left) / abo.budget)
        spent_color = gp.green
        unspent_color = gp.yellow
    bar_actual = bar_left + actual_width
    for i in range(bar_left, bar_right):
        cell = ws.cell(row = rowno, column = gp.COL_0 + i)
        cell.fill = spent_color if i < bar_actual else unspent_color
        if i == bar_left:
            cell.border = gp.black_tbl
        elif i == bar_right - 1:
            cell.border = gp.black_tbr
        elif i == curmonth_idx:
            cell.border = gp.black_tb_bluedash_r
        else:
            cell.border = gp.black_tb
    # Overrun cells
    bar_max = gp.horiz_ncells() + gp.overrun_ncells() - 1
    bar_right_rendered = bar_right
    if abo.budget and bar_actual > bar_right:
        if bar_actual <= bar_max:
            bar_ovr = bar_actual
            truncate = False
        else:
            bar_ovr = bar_max
            truncate = True # Render with darker red
            print("Warning: overrun exceeded, try OVERRUN_WIDTH =",
                  math.ceil((bar_actual - bar_right) / gp.month_hcells))
        for i in range(bar_right, bar_ovr):
            cell = ws.cell(row = rowno, column = gp.COL_0 + i)
            if i < bar_ovr - 1:
                cell.border = gp.black_tb
                cell.fill = gp.red1
            else:
                cell.border = gp.black_tbr
                cell.fill = gp.red2 if truncate else gp.red1
        bar_right_rendered = bar_ovr
    # Projected overrun cells
    bar_po = bar_right + po_width
    if bar_po > bar_actual: # Skip if actual exceeds projected overrun
        if bar_po <= bar_max:
            bar_ovr = bar_po
        else:
            bar_ovr = bar_max
            print("Warning: projected overrun exceeded, try OVERRUN_WIDTH =",
                  math.ceil((bar_po - bar_right) / gp.month_hcells))
        for i in range(bar_right_rendered, bar_ovr):
            cell = ws.cell(row = rowno, column = gp.COL_0 + i)
            cell.fill = gp.redshaded
            if i < bar_ovr - 1:
                cell.border = gp.reddash_tb
            else:
                cell.border = gp.reddash_tbr
    ws.row_dimensions[rowno].height = height
    # Percentage column
    if abo.budget:
        pct = ws.cell(row = rowno, column = gp.COL_PCT,
                      value = abo.actual / abo.budget)
        arial(pct)
        pct.number_format = "##0.0%"
        pct.alignment = Alignment(horizontal = 'right', vertical='center')
    # Actual/Budget in $k
    actbud = ws.cell(row = rowno, column = gp.COL_ACTBUD,
                     value = (str(round(abo.actual / 1000)) + "k/" +
                              str(round(abo.budget / 1000)) + "k"))
    arial(actbud)
    actbud.alignment = Alignment(horizontal = 'right', vertical='center')
    # Projected Overrun in $k
    if abo.proj_ovr:
        prjovr = ws.cell(row = rowno, column = gp.COL_PRJOVR,
                         value = f"[+{round(abo.proj_ovr / 1000)}k]")
        arial(prjovr, color="FF0000")
        prjovr.alignment = Alignment(horizontal = 'left', vertical='center')
    # Bucket name/descr
    bkt = ws.cell(row = rowno, column = gp.COL_DESCR, value = descr)
    arial(bkt)
    bkt.alignment = Alignment(horizontal = 'left', vertical='center',
                              indent=indent)

def create_buckets_worksheet(wb, gp, buckets):
    ws = wb.create_sheet(title = gp.sheetname)
    for colno in range(1, gp.COL_OVR):
        ws.column_dimensions[get_column_letter(colno)].width = gp.cell_width
    ws.column_dimensions[get_column_letter(gp.COL_OVR)].width = 2 # blank
    ws.column_dimensions[get_column_letter(gp.COL_PCT)].width = 6 # pct
    ws.column_dimensions[get_column_letter(gp.COL_ACTBUD)].width = 8 # act/bud
    ws.column_dimensions[get_column_letter(gp.COL_PRJOVR)].width = 5 # p.ovr
    ws.column_dimensions[get_column_letter(gp.COL_DESCR)].width = 30 # acct
    last_colno = gp.COL_DESCR
    rowno = gp.ROW_0
    for bucket_id in sorted(buckets):
        bd = buckets[bucket_id]
        render_parent = bool(bd.abo)
        if render_parent:
            descr = bd.name
            if len(bd.absorbed_accts) > 1:
                descr += f" ({len(bd.absorbed_accts)})"
            if bd.subbuckets:
                descr += " + ..."
            render_bucket(ws, rowno, gp, descr, bd.abo)
            rowno += 1
        for acct in sorted(bd.subbuckets):
            abo = bd.subbuckets[acct]
            if not abo:
                print("Warning: skipping 0 acct", acct)
                continue
            render_bucket(ws, rowno, gp, acct, abo, indent=render_parent)
            rowno += 1
    # Write headers and footers - do this after we know last_colno
    # Headers: title and months
    ttl = ws.cell(row=1, column=gp.COL_0, value="Expenses vs Budget")
    Style(fontsize=14, bold=True, center=True).text(ttl)
    merge_row(ws, 1, gp.COL_0, last_colno)
    ws.row_dimensions[1].height = 18
    if gp.months:
        monthstr = gp.months
        if gp.nmonths:
            monthstr += " (" + str(round(gp.nmonths * 100 / 12)) + "% of year)"
        mths = ws.cell(row=2, column=gp.COL_0, value=monthstr)
        Style(fontsize=12, bold=True, center=True).text(mths)
        merge_row(ws, 2, gp.COL_0, last_colno)
        ws.row_dimensions[2].height = 16

    # Footer (key)
    boxleft = gp.COL_0 + 2
    boxmid = boxleft + 3
    boxright = boxmid + 3
    textleft = boxright + 3
    textright = last_colno
    def keyrow(rowno, text, fill, border = gp.black_tblr):
        if fill is not None:
            box = ws.cell(row=rowno, column=boxleft)
            box.fill = fill
            box.border = border
            merge_row(ws, rowno, boxleft, boxright)
        txtcell = ws.cell(row=rowno, column=textleft, value=text)
        arial(txtcell)
        txtcell.alignment = Alignment(
            horizontal='left', vertical='top', wrap_text=True)
        merge_row(ws, rowno, textleft, textright)
        ws.row_dimensions[rowno].height = 12
    # Green
    rowno += 2
    keyrow(rowno, "Amount spent YTD", gp.green)
    rowno += 1
    keyrow(rowno, "Budgeted/unspent", gp.yellow)
    if gp.nmonths is not None:
        rowno += 1
        box = ws.cell(row=rowno, column=boxmid)
        box.border = gp.bluedash_r
        keyrow(rowno, f"Position in year (month {gp.nmonths} of 12)", None)
    rowno += 1
    keyrow(rowno, "Not budgeted", gp.cyan)
    rowno += 1
    keyrow(rowno, "Budget Overrun", gp.red1)
    rowno += 1
    keyrow(rowno, "Projected Overrun", gp.redshaded, gp.reddash_tblr)
    return ws


def process_expense_buckets(wb, a2e):
    gp = GraphParams("Buckets", a2e.months, a2e.nmonths)
    buckets = collect_expense_buckets(a2e, gp)
    ws = create_buckets_worksheet(wb, gp, buckets)
    # Random stats
    verbose("Budget total",
            round(sum([bd.abo.budget for bd in buckets.values()]), 2))
    verbose("Budget min",
            min([bd.abo.budget for bd in buckets.values()]))

class PNL(object):
    def __init__(self, fname, wb, ws):
        self.fname = fname
        self.wb = wb
        self.ws = ws
        self.out_fname = self._compute_out_fname() # does eager sanity assert

    def _compute_out_fname(self):
        tmp = self.fname.rsplit('.', 1)
        if len(tmp) == 2:
            assert tmp[1] in ("xls", "xlsx")
        return tmp[0] + "-mod.xlsx"

def main(args):
    budget_ws = None
    pnl = None
    for arg in args:
        wb = load_workbook(arg)
        wbl = len(wb.worksheets)
        if len(wb.worksheets) > 1:
            print(f"Warning: {arg} has {wbl} worksheets, doing first only")
        ws = wb.worksheets[0]
        a1 = ws.cell(1,1)
        if a1.value == "Profit and Loss":
            # Modern view:
            #  - column headers are "Distribution account" and "Total"
            #  - subaccounts are grouped
            assert pnl is None
            pnl = PNL(arg, wb, ws)
            continue
        a2 = ws.cell(2,1)
        if a2.value == "Profit and Loss":
            # Classic view:
            #  - column 1 header is empty, column 2 is "Total"
            #  - no grouping of subaccounts
            assert pnl is None
            pnl = PNL(arg, wb, ws)
            print(f"Warning: PnL might be classic view, use modern instead")
            sys.exit(1) # XXX - TODO, add support anyways?
            continue
        b1 = ws.cell(1,2)
        if b1.value == "Budget":
            assert budget_ws is None
            budget_ws = ws
            continue
        raise Err(f"Unrecognized spreadsheet {arg}")
    if not budget_ws:
        raise Err("Budget not loaded")
    if not pnl.ws:
        raise Err("P&L not loaded")
    a2e = Acct2Entries()
    read_entries(budget_ws, a2e)
    read_entries(pnl.ws, a2e, keep_affixes = True)
    wb = Workbook()
    addcells_pnl_vs_budget(wb.active, pnl.ws, a2e)
    set_column_widths(wb.active)
    process_expense_buckets(wb, a2e)
    wb.save(pnl.out_fname)
    print("Wrote", pnl.out_fname)

if __name__ == "__main__":
    main(sys.argv[1:])
