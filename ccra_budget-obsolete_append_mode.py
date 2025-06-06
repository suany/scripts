#!/usr/bin/env python3
#
# OBSOLETE APPEND MODE
# This version simply tacks on budget column to a P&L.
# It does not insert budget rows with no expenses.
#
# openpyxl or xlswriter
#
# Alt: Data processing in python:
#   pandas library: pd.read_excel()
#   DataFrame.plot
#
# Excel functions
#   xlookup (2021), mine is 2016
#   -> in any case, would be quite clunky

import sys
from copy import copy
from openpyxl import load_workbook # pip3 install openpyxl

class Err(Exception):
    def __str__(self):
        return "\n" + self.args[0]

def verbose(*args, **kwargs):
    pass
    #print(*args, **kwargs)

def set_column_widths(ws):
    # Set column widths - these work well for 2025
    ws.column_dimensions['A'].width = 31 # =310px
    ws.column_dimensions['B'].width = 9
    ws.column_dimensions['C'].width = 9
    ws.column_dimensions['D'].width = 7
    ws.column_dimensions['E'].width = 1
    ws.column_dimensions['F'].width = 26

def read_budget_and_notes(budget):
    budget_dict = dict()
    notes_dict = dict()
    first_time = True # to check header
    for row in budget.rows:
        assert len(row) >= 3
        acct = row[0].value
        amt = row[1].value
        notes = row[2].value
        if first_time:
            assert acct == "Distribution account"
            assert amt == "Budget"
            assert notes == "Notes"
            first_time = False
            # Insert into table anyways
        if not acct:
            verbose("Empty first cell, stopping at", row)
            break
        if acct in budget_dict:
            print("Warning: repeated budget key(", acct, ") at", row)
        if acct in notes_dict:
            print("Warning: repeated notes key(", acct, ") at", row)
        if amt:
            budget_dict[acct] = amt
        if notes:
            notes_dict[acct] = notes
    return budget_dict, notes_dict

def copy_styles(src, dst,
                font = None,
                border = None,
                fill = None,
                number_format = None,
                protection = None,
                alignment = None):
    if src.has_style:
        dst.font = (copy(src.font)
                    if font is None else font)
        dst.border = (copy(src.border)
                      if border is None else border)
        dst.fill = (copy(src.fill)
                    if fill is None else fill)
        dst.number_format = (copy(src.number_format)
                             if number_format is None else number_format)
        dst.protection = (copy(src.protection)
                          if protection is None else protection)
        dst.alignment = (copy(src.alignment)
                         if alignment is None else alignment)

def append_budget_and_notes(pnlws, budget_dict, notes_dict):
    for row in pnlws.rows:
        assert len(row) <= 2
        acct = row[0].value
        rowno = row[0].row
        budget = budget_dict.get(acct)
        notes = notes_dict.get(acct)
        if budget:
            c3 = pnlws.cell(row=rowno, column=3, value=budget)
            copy_styles(row[1], c3)
            if budget == "Budget": # Header row
                c4 = pnlws.cell(row=rowno, column=4, value="Pct")
                copy_styles(row[1], c4)
            else:
                c4 = pnlws.cell(row=rowno, column=4,
                                value=f"=B{rowno}/C{rowno}")
                copy_styles(row[1], c4, number_format = "##0.0%")
        if notes:
            c6 = pnlws.cell(row=rowno, column=6, value=notes)
            copy_styles(row[0], c6)

def expand_merge(pnlws, mcr):
    if mcr.min_row == mcr.max_row and mcr.min_col == 1 and mcr.max_col == 2:
        pnlws.unmerge_cells(
            start_row=mcr.min_row,
            start_column=mcr.min_col,
            end_row=mcr.max_row,
            end_column=mcr.max_col)
        pnlws.merge_cells(
            start_row=mcr.min_row,
            start_column=mcr.min_col,
            end_row=mcr.max_row,
            end_column=mcr.max_col + 4)
    else:
        print("Warning: expand merge skipping", mcr)

def reformat_updated_pnl(pnlws):
    set_column_widths(pnlws)
    # TODO: set row heights for multilines?
    # Expand merged cells
    for mcr in list(pnlws.merged_cells.ranges):
        expand_merge(pnlws, mcr)

class PNL(object):
    def __init__(self, fname, wb, ws):
        self.fname = fname
        self.wb = wb
        self.ws = ws
        self.out_fname = self._compute_out_fname() # does eager sanity assert

    def _compute_out_fname(self):
        tmp = self.fname.rsplit('.', 1)
        if len(tmp) == 2:
            assert tmp[1] in ("xls", "xlsx")
        return tmp[0] + "-mod1.xlsx"

def main(args):
    budget_ws = None
    pnl = None
    for arg in args:
        wb = load_workbook(arg)
        wbl = len(wb.worksheets)
        if len(wb.worksheets) > 1:
            print(f"Warning: {arg} has {wbl} worksheets, doing first only")
        ws = wb.worksheets[0]
        a1 = ws.cell(1,1)
        if a1.value == "Profit and Loss":
            assert pnl is None
            pnl = PNL(arg, wb, ws)
            continue
        b1 = ws.cell(1,2)
        if b1.value == "Budget":
            assert budget_ws is None
            budget_ws = ws
            continue
        raise Err("Unrecognized spreadsheet")
    if not budget_ws:
        raise Err("Budget not loaded")
    if not pnl.ws:
        raise Err("P&L not loaded")
    budget_dict, notes_dict = read_budget_and_notes(budget_ws)
    append_budget_and_notes(pnl.ws, budget_dict, notes_dict)
    reformat_updated_pnl(pnl.ws)
    pnl.wb.save(pnl.out_fname)

if __name__ == "__main__":
    main(sys.argv[1:])
